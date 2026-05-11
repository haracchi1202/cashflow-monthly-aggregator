"""資金繰り予算ファイルの解析・上書きプラン・上書き実行。

予算ファイル想定レイアウト:

    A             B    C    D    E    F    G    ...
                       4月       5月       6月       ...
                       予算 実績 予算 実績 予算 実績 ...
    確定入金 ...
    確定支払 ...
    予測入金 ...   ← 任意（存在すれば検出して書き込み対象に追加）
    予測支払 ...   ← 任意

[安全設計]
    - 原本は触らず、メモリ上のコピーに対して上書き
    - 上書きするのは事前検出した予算列セル（確定入金/確定支払/予測入金/予測支払 × 各月）のみ
    - 確定値は確定行、予測値は予測行に**別々に**書き込む（合算しない）
    - 予測行が存在しない場合は確定行のみ更新（警告を出す）
    - 実績列・他シート・他行・数式・書式は一切変更しない
    - 上書き前後確認を必ず通す
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from excel_reader import normalize_month


DEFAULT_TARGET_SHEETS = ("MBJ上半期", "MBJ下半期")


# =========================================================
# データクラス
# =========================================================
@dataclass
class BudgetCellMap:
    """予算ファイル内で「上書き候補」となるセル1つを表す。"""
    sheet: str
    month: str                  # YYYY-MM
    transaction_status: str     # "confirmed" or "forecast"
    kind: str                   # "入金" or "出金"
    row: int
    col: int
    cell_address: str           # 例: "F32"
    before_value: float | None
    before_raw: Any


@dataclass
class SheetDiagnostics:
    sheet: str
    found: bool
    month_header_row: int | None = None
    subheader_row: int | None = None
    confirmed_income_row: int | None = None
    confirmed_expense_row: int | None = None
    forecast_income_row: int | None = None
    forecast_expense_row: int | None = None
    months: list[str] = field(default_factory=list)
    error: str | None = None
    notes: list[str] = field(default_factory=list)


@dataclass
class CellPlan:
    """1セル分の更新プラン。月×シート×ステータス×種別 で 1 行。"""
    sheet: str
    month: str
    transaction_status: str     # "confirmed" or "forecast"
    kind: str                   # "入金" or "出金"
    cell_address: str
    before: float | None
    after: float | None
    diff: float | None
    in_range: bool
    has_aggregate: bool
    update_flag: bool
    reason: str


# =========================================================
# 補助
# =========================================================
def _cell_text(cell) -> str:
    val = cell.value
    if val is None:
        return ""
    return str(val).strip()


def _norm(text: str) -> str:
    return text.replace(" ", "").replace("　", "")


def _merged_range_for(ws, row: int, col: int) -> tuple[int, int, int, int]:
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            return mr.min_row, mr.min_col, mr.max_row, mr.max_col
    return row, col, row, col


def _infer_fiscal_year(sheet_name: str, default_year: int | None) -> int | None:
    m = re.search(r"(?:FY|fy)\s*(\d{2,4})", sheet_name)
    if not m:
        m = re.search(r"(\d{2,4})\s*年", sheet_name)
    if m:
        y = int(m.group(1))
        if y < 100:
            y += 2000
        return y
    return default_year


def _is_first_half(s: str) -> bool:
    return "上半期" in s or "上期" in s


def _is_second_half(s: str) -> bool:
    return "下半期" in s or "下期" in s


def _resolve_year_for_month(sheet_name: str, month_num: int, fiscal_year: int) -> int:
    if _is_first_half(sheet_name):
        return fiscal_year
    if _is_second_half(sheet_name):
        if 10 <= month_num <= 12:
            return fiscal_year
        if 1 <= month_num <= 3:
            return fiscal_year + 1
    return fiscal_year


# =========================================================
# 検出ロジック
# =========================================================
def _find_subheader_row(ws) -> int | None:
    best_row: int | None = None
    best_count = 0
    max_row = min(ws.max_row, 200)
    max_col = min(ws.max_column, 200)
    for row in range(1, max_row + 1):
        budget = 0
        actual = 0
        for col in range(1, max_col + 1):
            text = _norm(_cell_text(ws.cell(row, col)))
            if not text:
                continue
            if text in ("予算", "予算額"):
                budget += 1
            elif text in ("実績", "実績額"):
                actual += 1
        count = min(budget, actual)
        if count > best_count:
            best_count = count
            best_row = row
    return best_row if best_count >= 2 else None


def _detect_month_headers(
    ws,
    sheet_name: str,
    subheader_row: int,
    fiscal_year: int | None,
) -> tuple[dict[str, int], int | None]:
    max_col = min(ws.max_column, 200)
    candidate_rows = list(range(max(1, subheader_row - 6), subheader_row))
    for header_row in reversed(candidate_rows):
        month_to_col: dict[str, int] = {}
        cols_done: set[int] = set()
        for col in range(1, max_col + 1):
            if col in cols_done:
                continue
            cell = ws.cell(header_row, col)
            text = _cell_text(cell)
            if not text:
                continue
            ym = normalize_month(text)
            if not ym:
                m = re.match(r"^\s*(\d{1,2})\s*月\s*$", text)
                if m and fiscal_year is not None:
                    month_num = int(m.group(1))
                    year = _resolve_year_for_month(sheet_name, month_num, fiscal_year)
                    ym = f"{year:04d}-{month_num:02d}"
            if not ym:
                continue
            _, c1, _, c2 = _merged_range_for(ws, header_row, col)
            for cc in range(c1, c2 + 1):
                cols_done.add(cc)
            budget_col: int | None = None
            for cc in range(c1, c2 + 1):
                sub_text = _norm(_cell_text(ws.cell(subheader_row, cc)))
                if "予算" in sub_text and "実績" not in sub_text:
                    budget_col = cc
                    break
            if budget_col is not None:
                month_to_col[ym] = budget_col
        if len(month_to_col) >= 2:
            return month_to_col, header_row
    return {}, None


# 行ラベルパターン
_CONFIRMED_INCOME_PAT = re.compile(r"確定.{0,2}入金")
_CONFIRMED_EXPENSE_PAT = re.compile(r"確定.{0,2}支払")
_FORECAST_INCOME_PATS = [
    re.compile(r"予測.{0,2}入金"),
    re.compile(r"見込.{0,2}入金"),
    re.compile(r"予定.{0,2}入金"),
]
_FORECAST_EXPENSE_PATS = [
    re.compile(r"予測.{0,2}支払"),
    re.compile(r"見込.{0,2}支払"),
    re.compile(r"予定.{0,2}支払"),
]


def _find_kind_rows(ws) -> tuple[int | None, int | None, int | None, int | None]:
    """確定入金 / 確定支払 / 予測入金 / 予測支払 行を検出。

    返り値: (confirmed_income, confirmed_expense, forecast_income, forecast_expense)
    """
    confirmed_income: int | None = None
    confirmed_expense: int | None = None
    forecast_income: int | None = None
    forecast_expense: int | None = None
    max_row = min(ws.max_row, 400)
    label_cols = min(ws.max_column, 8)
    for row in range(1, max_row + 1):
        for col in range(1, label_cols + 1):
            text = _norm(_cell_text(ws.cell(row, col)))
            if not text:
                continue

            # 確定/予測 のどちらにマッチするかを先に判定（先勝ち）
            is_forecast_income = any(p.search(text) for p in _FORECAST_INCOME_PATS)
            is_forecast_expense = any(p.search(text) for p in _FORECAST_EXPENSE_PATS)
            is_confirmed_income = (
                not is_forecast_income
                and _CONFIRMED_INCOME_PAT.search(text) is not None
                and "支払" not in text
            )
            is_confirmed_expense = (
                not is_forecast_expense
                and _CONFIRMED_EXPENSE_PAT.search(text) is not None
            )

            if confirmed_income is None and is_confirmed_income:
                confirmed_income = row
            if confirmed_expense is None and is_confirmed_expense:
                confirmed_expense = row
            if forecast_income is None and is_forecast_income:
                forecast_income = row
            if forecast_expense is None and is_forecast_expense:
                forecast_expense = row
        if confirmed_income and confirmed_expense and forecast_income and forecast_expense:
            break
    return confirmed_income, confirmed_expense, forecast_income, forecast_expense


# =========================================================
# 解析エントリ
# =========================================================
def analyze_budget_workbook(
    file_bytes: bytes,
    target_sheets: tuple[str, ...] = DEFAULT_TARGET_SHEETS,
    fiscal_year_override: int | None = None,
) -> tuple[list[BudgetCellMap], list[SheetDiagnostics], list[str]]:
    cell_maps: list[BudgetCellMap] = []
    diagnostics: list[SheetDiagnostics] = []
    log_lines: list[str] = []

    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=False, keep_vba=False)
        wb_v = load_workbook(io.BytesIO(file_bytes), data_only=True, keep_vba=False)
    except Exception as e:
        log_lines.append(f"予算ファイル読み込み失敗: {e}")
        return [], [SheetDiagnostics(sheet="(全体)", found=False, error=str(e))], log_lines

    for sheet_name in target_sheets:
        diag = SheetDiagnostics(sheet=sheet_name, found=False)
        if sheet_name not in wb.sheetnames:
            diag.error = "シートが見つかりません"
            diagnostics.append(diag)
            log_lines.append(f"[{sheet_name}] シートが見つかりません")
            continue

        ws = wb[sheet_name]
        ws_v = wb_v[sheet_name]

        sub_row = _find_subheader_row(ws)
        if sub_row is None:
            diag.error = "『予算/実績』サブヘッダ行が見つかりません"
            diagnostics.append(diag)
            log_lines.append(f"[{sheet_name}] 『予算/実績』サブヘッダ未検出")
            continue
        diag.subheader_row = sub_row

        fy = _infer_fiscal_year(sheet_name, fiscal_year_override)
        if fy is None:
            diag.notes.append("年度が推測できないため、月のみ表記の月は無視されます")

        month_to_col, month_row = _detect_month_headers(ws, sheet_name, sub_row, fy)
        if not month_to_col:
            diag.error = "月ヘッダから『予算』列を特定できません"
            diagnostics.append(diag)
            continue
        diag.month_header_row = month_row
        diag.months = sorted(month_to_col.keys())

        ci_row, ce_row, fi_row, fe_row = _find_kind_rows(ws)
        if ci_row is None or ce_row is None:
            diag.error = (
                f"確定入金/確定支払 行を特定できません "
                f"(confirmed_income={ci_row}, confirmed_expense={ce_row})"
            )
            diagnostics.append(diag)
            continue
        diag.confirmed_income_row = ci_row
        diag.confirmed_expense_row = ce_row
        diag.forecast_income_row = fi_row
        diag.forecast_expense_row = fe_row

        if fi_row is None and fe_row is None:
            diag.notes.append(
                "予測入金/予測支払 行が見つかりません。予測値の書き込みはスキップされます。"
            )
        elif fi_row is None:
            diag.notes.append("予測入金 行のみ見つかりませんでした。")
        elif fe_row is None:
            diag.notes.append("予測支払 行のみ見つかりませんでした。")

        row_table: list[tuple[str, str, int | None]] = [
            ("confirmed", "入金", ci_row),
            ("confirmed", "出金", ce_row),
            ("forecast", "入金", fi_row),
            ("forecast", "出金", fe_row),
        ]
        for ym, col in sorted(month_to_col.items()):
            for status, kind, r in row_table:
                if r is None:
                    continue
                v_cell = ws_v.cell(r, col)
                f_cell = ws.cell(r, col)
                value = v_cell.value
                before_num: float | None
                if isinstance(value, (int, float)) and value == value:
                    before_num = float(value)
                else:
                    before_num = None
                cell_maps.append(
                    BudgetCellMap(
                        sheet=sheet_name,
                        month=ym,
                        transaction_status=status,
                        kind=kind,
                        row=r,
                        col=col,
                        cell_address=f"{get_column_letter(col)}{r}",
                        before_value=before_num,
                        before_raw=f_cell.value,
                    )
                )

        diag.found = True
        diagnostics.append(diag)
        log_lines.append(
            f"[{sheet_name}] 解析成功: 月数={len(month_to_col)}, "
            f"確定入金行={ci_row}, 確定支払行={ce_row}, "
            f"予測入金行={fi_row}, 予測支払行={fe_row}, "
            f"月ヘッダ行={month_row}, サブヘッダ行={sub_row}"
        )

    return cell_maps, diagnostics, log_lines


# =========================================================
# 上書きプラン
# =========================================================
UpdateScope = str  # "both" | "confirmed_only" | "forecast_only"


def build_overwrite_plan(
    cell_maps: list[BudgetCellMap],
    confirmed_summary: dict[str, dict[str, float]],
    forecast_summary: dict[str, dict[str, float]],
    month_from: str | None,
    month_to: str | None,
    update_scope: UpdateScope = "both",
) -> tuple[list[CellPlan], list[str], list[str]]:
    """セル単位の更新プランを返す。

    Parameters
    ----------
    cell_maps
        analyze_budget_workbook で得た上書き候補セル一覧
    confirmed_summary / forecast_summary
        それぞれ {YYYY-MM: {"入金合計": x, "出金合計": y}}
    update_scope
        "both" / "confirmed_only" / "forecast_only"

    Returns
    -------
    plans
        セル単位の更新判定一覧
    summary_only_months
        集計データにあるが予算ファイルにない月
    budget_only_months
        予算ファイルにあるが集計データにない月（指定期間内のみ）
    """
    plans: list[CellPlan] = []
    for cm in cell_maps:
        in_range = True
        if month_from and cm.month < month_from:
            in_range = False
        if month_to and cm.month > month_to:
            in_range = False

        # スコープ判定
        scope_ok = (
            update_scope == "both"
            or (update_scope == "confirmed_only" and cm.transaction_status == "confirmed")
            or (update_scope == "forecast_only" and cm.transaction_status == "forecast")
        )

        if cm.transaction_status == "confirmed":
            agg = confirmed_summary.get(cm.month)
        else:
            agg = forecast_summary.get(cm.month)
        has_aggregate = agg is not None

        if not in_range:
            reason = "指定期間外"
            update = False
            after = cm.before_value
        elif not scope_ok:
            reason = "対象外（スコープ）"
            update = False
            after = cm.before_value
        elif not has_aggregate:
            reason = "集計データなし"
            update = False
            after = cm.before_value
        else:
            reason = "正常更新"
            update = True
            key = "入金合計" if cm.kind == "入金" else "出金合計"
            after = float(agg.get(key, 0) or 0)

        diff = (
            (float(after) - float(cm.before_value))
            if (after is not None and cm.before_value is not None)
            else None
        )

        plans.append(
            CellPlan(
                sheet=cm.sheet,
                month=cm.month,
                transaction_status=cm.transaction_status,
                kind=cm.kind,
                cell_address=cm.cell_address,
                before=cm.before_value,
                after=after,
                diff=diff,
                in_range=in_range,
                has_aggregate=has_aggregate,
                update_flag=update,
                reason=reason,
            )
        )

    # 未照合月
    budget_months = {cm.month for cm in cell_maps}
    summary_months = set(confirmed_summary.keys()) | set(forecast_summary.keys())
    summary_only = sorted(summary_months - budget_months)
    budget_only: list[str] = []
    for m in sorted(budget_months - summary_months):
        if month_from and m < month_from:
            continue
        if month_to and m > month_to:
            continue
        budget_only.append(m)
    return plans, summary_only, budget_only


# =========================================================
# 上書き実行 / 検証
# =========================================================
def apply_overwrite(
    original_bytes: bytes,
    plans: list[CellPlan],
    cell_maps: list[BudgetCellMap],
) -> tuple[bytes, int, int, list[dict[str, Any]]]:
    """事前生成したプランを元に上書き実行。

    プラン側で update_flag=True のセルだけ書き込む。
    """
    # cell_address & sheet で BudgetCellMap を引けるようにする
    by_addr: dict[tuple[str, str], BudgetCellMap] = {
        (cm.sheet, cm.cell_address): cm for cm in cell_maps
    }

    wb = load_workbook(io.BytesIO(original_bytes), data_only=False, keep_vba=False)
    updated_count = 0
    skipped_count = 0
    written_log: list[dict[str, Any]] = []

    for plan in plans:
        if not plan.update_flag:
            skipped_count += 1
            continue
        cm = by_addr.get((plan.sheet, plan.cell_address))
        if cm is None:
            skipped_count += 1
            continue
        ws = wb[cm.sheet]
        before = ws.cell(cm.row, cm.col).value
        new_value = float(plan.after) if plan.after is not None else None
        ws.cell(cm.row, cm.col).value = new_value
        updated_count += 1
        written_log.append(
            {
                "sheet": cm.sheet,
                "month": cm.month,
                "transaction_status": cm.transaction_status,
                "kind": cm.kind,
                "cell": cm.cell_address,
                "before": before,
                "after": new_value,
            }
        )

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read(), updated_count, skipped_count, written_log


def verify_overwrite(
    updated_bytes: bytes,
    cell_maps: list[BudgetCellMap],
) -> list[dict[str, Any]]:
    wb = load_workbook(io.BytesIO(updated_bytes), data_only=False, keep_vba=False)
    out: list[dict[str, Any]] = []
    for cm in cell_maps:
        ws = wb[cm.sheet]
        after = ws.cell(cm.row, cm.col).value
        out.append(
            {
                "sheet": cm.sheet,
                "month": cm.month,
                "transaction_status": cm.transaction_status,
                "kind": cm.kind,
                "cell": cm.cell_address,
                "after": after,
            }
        )
    return out
