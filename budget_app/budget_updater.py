"""資金繰り予算ファイルの解析・上書きプラン作成・実行。

予算ファイル (MBJ26年資金繰...xlsx) は次のレイアウトを想定:

    A             B    C    D    E    F    G    ...
                       4月       5月       6月       ...
                       予算 実績 予算 実績 予算 実績 ...
    確定入金                                              <-- 確定入金行
    確定支払                                              <-- 確定支払行

月ヘッダはセル結合されている可能性があるため、結合範囲ごとに「予算」列を探す。
シート名に「26年」「上半期 / 下半期」が含まれることを利用して、月のみの表記
（"4月" など）からも YYYY-MM を推測する。

安全設計:
  - 原本は触らず、in-memory コピーに対して上書きする
  - 上書きするのは事前検出した (確定入金行|確定支払行) × 予算列 セルのみ
  - 実績列・他シート・他行は一切変更しない
  - 数式・書式は openpyxl の load_workbook(data_only=False) を使うため自動的に保持
  - 上書き前後の値を必ず計算して比較レポートに出す
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from excel_reader import normalize_month


DEFAULT_TARGET_SHEETS = ("MBJ上半期", "MBJ下半期")


# =========================================================
# データクラス
# =========================================================
@dataclass
class BudgetCellMap:
    """予算ファイル内で「上書き候補」となるセル1つを表す。"""
    sheet: str
    month: str          # YYYY-MM
    kind: str           # "入金" or "出金"
    row: int            # 1始まり
    col: int            # 1始まり
    cell_address: str   # 例: "F32"
    before_value: float | None
    before_raw: Any     # 数式文字列の可能性あり


@dataclass
class SheetDiagnostics:
    sheet: str
    found: bool
    month_header_row: int | None = None
    subheader_row: int | None = None
    income_row: int | None = None
    expense_row: int | None = None
    months: list[str] = field(default_factory=list)
    error: str | None = None
    notes: list[str] = field(default_factory=list)


# =========================================================
# 補助
# =========================================================
def _cell_text(cell) -> str:
    val = cell.value
    if val is None:
        return ""
    return str(val).strip()


def _norm(text: str) -> str:
    return text.replace(" ", "").replace("　", "")


def _merged_range_for(ws, row: int, col: int) -> tuple[int, int, int, int]:
    """指定セルが含まれる結合範囲を (r1, c1, r2, c2) で返す。結合外ならそのセル単体。"""
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            return mr.min_row, mr.min_col, mr.max_row, mr.max_col
    return row, col, row, col


def _infer_fiscal_year(sheet_name: str, default_year: int | None) -> int | None:
    """シート名から年度を推測。'MBJ26年...' → 2026。'FY26' → 2026。見つからなければ default_year。"""
    m = re.search(r"(?:FY|fy|エフワイ)\s*(\d{2,4})", sheet_name)
    if not m:
        m = re.search(r"(\d{2,4})\s*年", sheet_name)
    if m:
        y = int(m.group(1))
        if y < 100:
            y += 2000
        return y
    return default_year


def _is_first_half(sheet_name: str) -> bool:
    return "上半期" in sheet_name or "上期" in sheet_name


def _is_second_half(sheet_name: str) -> bool:
    return "下半期" in sheet_name or "下期" in sheet_name


def _resolve_year_for_month(sheet_name: str, month_num: int, fiscal_year: int) -> int:
    """日本の会計年度（4月始まり）を仮定し、半期と月から実年を返す。"""
    if _is_first_half(sheet_name):
        return fiscal_year
    if _is_second_half(sheet_name):
        if 10 <= month_num <= 12:
            return fiscal_year
        if 1 <= month_num <= 3:
            return fiscal_year + 1
    return fiscal_year


# =========================================================
# 検出ロジック
# =========================================================
def _find_budget_actual_subheader_row(ws) -> int | None:
    """『予算/実績』のペアが最も多く並んでいる行を返す。"""
    best_row: int | None = None
    best_count = 0
    max_row = min(ws.max_row, 200)
    max_col = min(ws.max_column, 200)
    for row in range(1, max_row + 1):
        budget = 0
        actual = 0
        for col in range(1, max_col + 1):
            text = _norm(_cell_text(ws.cell(row, col)))
            if not text:
                continue
            if text in ("予算", "予算額"):
                budget += 1
            elif text in ("実績", "実績額"):
                actual += 1
        count = min(budget, actual)
        if count > best_count:
            best_count = count
            best_row = row
    return best_row if best_count >= 2 else None


def _detect_month_headers(
    ws,
    sheet_name: str,
    subheader_row: int,
    fiscal_year: int | None,
) -> tuple[dict[str, int], int | None, list[str]]:
    """月ヘッダ行を検出し、各月の『予算』列を返す。

    Returns:
        ({YYYY-MM: 予算列番号}, 月ヘッダ行, ログメッセージ)
    """
    logs: list[str] = []
    max_col = min(ws.max_column, 200)

    # サブヘッダの上 6 行までを候補として走査
    candidate_rows = list(range(max(1, subheader_row - 6), subheader_row))
    for header_row in reversed(candidate_rows):
        month_to_col: dict[str, int] = {}
        cols_done: set[int] = set()
        for col in range(1, max_col + 1):
            if col in cols_done:
                continue
            cell = ws.cell(header_row, col)
            text = _cell_text(cell)
            if not text:
                continue

            ym = normalize_month(text)
            if not ym:
                m = re.match(r"^\s*(\d{1,2})\s*月\s*$", text)
                if m and fiscal_year is not None:
                    month_num = int(m.group(1))
                    year = _resolve_year_for_month(sheet_name, month_num, fiscal_year)
                    ym = f"{year:04d}-{month_num:02d}"

            if not ym:
                continue

            r1, c1, r2, c2 = _merged_range_for(ws, header_row, col)
            for cc in range(c1, c2 + 1):
                cols_done.add(cc)

            budget_col: int | None = None
            for cc in range(c1, c2 + 1):
                sub_text = _norm(_cell_text(ws.cell(subheader_row, cc)))
                if "予算" in sub_text and "実績" not in sub_text:
                    budget_col = cc
                    break
            if budget_col is not None:
                month_to_col[ym] = budget_col

        if len(month_to_col) >= 2:
            logs.append(f"月ヘッダ行={header_row}, 検出月数={len(month_to_col)}")
            return month_to_col, header_row, logs

    return {}, None, logs


def _find_kind_rows(ws) -> tuple[int | None, int | None]:
    """確定入金 / 確定支払 行を検出。フリガナ 'カクテイニュウキン' などが付いていても OK。"""
    income_row: int | None = None
    expense_row: int | None = None
    max_row = min(ws.max_row, 300)
    label_cols = min(ws.max_column, 8)
    for row in range(1, max_row + 1):
        for col in range(1, label_cols + 1):
            text = _norm(_cell_text(ws.cell(row, col)))
            if not text:
                continue
            if income_row is None and "確定入金" in text and "支払" not in text:
                income_row = row
            if expense_row is None and ("確定支払" in text or "確定支払い" in text):
                expense_row = row
        if income_row and expense_row:
            break
    return income_row, expense_row


# =========================================================
# 解析メインエントリ
# =========================================================
def analyze_budget_workbook(
    file_bytes: bytes,
    target_sheets: tuple[str, ...] = DEFAULT_TARGET_SHEETS,
    fiscal_year_override: int | None = None,
) -> tuple[list[BudgetCellMap], list[SheetDiagnostics], list[str]]:
    """予算ファイルを解析し、上書き候補セル一覧 + 各シートの診断 + ログを返す。

    Returns
    -------
    cell_maps
        全シート横断の上書き候補セル一覧
    diagnostics
        シートごとの検出結果（成功 / 失敗・検出位置）
    log_lines
        画面ログ用のテキスト
    """
    cell_maps: list[BudgetCellMap] = []
    diagnostics: list[SheetDiagnostics] = []
    log_lines: list[str] = []

    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=False, keep_vba=False)
        wb_v = load_workbook(io.BytesIO(file_bytes), data_only=True, keep_vba=False)
    except Exception as e:
        log_lines.append(f"予算ファイル読み込み失敗: {e}")
        return [], [SheetDiagnostics(sheet="(全体)", found=False, error=str(e))], log_lines

    for sheet_name in target_sheets:
        diag = SheetDiagnostics(sheet=sheet_name, found=False)
        if sheet_name not in wb.sheetnames:
            diag.error = "シートが見つかりません"
            diagnostics.append(diag)
            log_lines.append(f"[{sheet_name}] シートが見つかりません")
            continue

        ws = wb[sheet_name]
        ws_v = wb_v[sheet_name]

        sub_row = _find_budget_actual_subheader_row(ws)
        if sub_row is None:
            diag.error = "『予算/実績』サブヘッダ行が見つかりません"
            diagnostics.append(diag)
            log_lines.append(f"[{sheet_name}] 『予算/実績』サブヘッダ行を検出できず")
            continue
        diag.subheader_row = sub_row

        fy = _infer_fiscal_year(sheet_name, fiscal_year_override)
        if fy is None:
            diag.notes.append("年度が推測できないため、月のみ表記の月は無視されます")

        month_to_col, month_row, mlogs = _detect_month_headers(ws, sheet_name, sub_row, fy)
        log_lines.extend(f"[{sheet_name}] {m}" for m in mlogs)
        if not month_to_col:
            diag.error = "月ヘッダから『予算』列を特定できません"
            diagnostics.append(diag)
            continue
        diag.month_header_row = month_row
        diag.months = sorted(month_to_col.keys())

        income_row, expense_row = _find_kind_rows(ws)
        if income_row is None or expense_row is None:
            diag.error = (
                f"確定入金/確定支払 行を特定できません "
                f"(income={income_row}, expense={expense_row})"
            )
            diagnostics.append(diag)
            continue
        diag.income_row = income_row
        diag.expense_row = expense_row

        for ym, col in sorted(month_to_col.items()):
            for kind, r in (("入金", income_row), ("出金", expense_row)):
                v_cell = ws_v.cell(r, col)
                f_cell = ws.cell(r, col)
                value = v_cell.value
                before_num: float | None
                if isinstance(value, (int, float)) and value == value:  # nan check
                    before_num = float(value)
                else:
                    before_num = None
                cell_maps.append(
                    BudgetCellMap(
                        sheet=sheet_name,
                        month=ym,
                        kind=kind,
                        row=r,
                        col=col,
                        cell_address=f"{get_column_letter(col)}{r}",
                        before_value=before_num,
                        before_raw=f_cell.value,
                    )
                )

        diag.found = True
        diagnostics.append(diag)
        log_lines.append(
            f"[{sheet_name}] 解析成功: 月数={len(month_to_col)}, "
            f"確定入金行={income_row}, 確定支払行={expense_row}, "
            f"月ヘッダ行={month_row}, サブヘッダ行={sub_row}"
        )

    return cell_maps, diagnostics, log_lines


# =========================================================
# 上書きプラン
# =========================================================
@dataclass
class PlanRow:
    sheet: str
    month: str
    income_cell: str
    expense_cell: str
    income_before: float | None
    income_after: float | None
    income_diff: float | None
    expense_before: float | None
    expense_after: float | None
    expense_diff: float | None
    in_range: bool
    has_aggregate: bool
    update_flag: bool
    reason: str


def build_overwrite_plan(
    cell_maps: list[BudgetCellMap],
    monthly_summary: dict[str, dict[str, float]],
    month_from: str | None,
    month_to: str | None,
) -> tuple[list[PlanRow], list[str], list[str]]:
    """更新プランを生成。

    Returns
    -------
    plan_rows
        各 (シート, 月) ごとの更新判定行
    summary_only_months
        集計データにあるが予算ファイルにない月
    budget_only_months
        予算ファイルにあるが集計データにない月（指定期間内のもののみ）
    """
    by_sheet_month: dict[tuple[str, str], dict[str, BudgetCellMap]] = {}
    for cm in cell_maps:
        by_sheet_month.setdefault((cm.sheet, cm.month), {})[cm.kind] = cm

    plan_rows: list[PlanRow] = []
    for (sheet, month), kinds in sorted(by_sheet_month.items(), key=lambda x: (x[0][0], x[0][1])):
        in_range = True
        if month_from and month < month_from:
            in_range = False
        if month_to and month > month_to:
            in_range = False

        agg = monthly_summary.get(month)
        has_aggregate = agg is not None

        income_cm = kinds.get("入金")
        expense_cm = kinds.get("出金")
        income_before = income_cm.before_value if income_cm else None
        expense_before = expense_cm.before_value if expense_cm else None
        income_addr = income_cm.cell_address if income_cm else ""
        expense_addr = expense_cm.cell_address if expense_cm else ""

        if not in_range:
            reason = "指定期間外"
            update = False
            income_after = income_before
            expense_after = expense_before
        elif not has_aggregate:
            reason = "集計データなし"
            update = False
            income_after = income_before
            expense_after = expense_before
        else:
            reason = "正常更新"
            update = True
            income_after = float(agg.get("入金合計", 0) or 0) if income_cm else None
            expense_after = float(agg.get("出金合計", 0) or 0) if expense_cm else None

        def _diff(a: float | None, b: float | None) -> float | None:
            if a is None or b is None:
                return None
            return float(a) - float(b)

        plan_rows.append(
            PlanRow(
                sheet=sheet,
                month=month,
                income_cell=income_addr,
                expense_cell=expense_addr,
                income_before=income_before,
                income_after=income_after,
                income_diff=_diff(income_after, income_before),
                expense_before=expense_before,
                expense_after=expense_after,
                expense_diff=_diff(expense_after, expense_before),
                in_range=in_range,
                has_aggregate=has_aggregate,
                update_flag=update,
                reason=reason,
            )
        )

    budget_months = {month for (_, month) in by_sheet_month.keys()}
    summary_months = set(monthly_summary.keys())
    summary_only = sorted(summary_months - budget_months)

    budget_only: list[str] = []
    for m in sorted(budget_months - summary_months):
        if month_from and m < month_from:
            continue
        if month_to and m > month_to:
            continue
        budget_only.append(m)

    return plan_rows, summary_only, budget_only


# =========================================================
# 実書き込み
# =========================================================
def apply_overwrite(
    original_bytes: bytes,
    cell_maps: list[BudgetCellMap],
    monthly_summary: dict[str, dict[str, float]],
    month_from: str | None,
    month_to: str | None,
) -> tuple[bytes, int, int, list[dict[str, Any]]]:
    """原本のコピーに上書きして bytes・更新件数・スキップ件数・ログを返す。

    実績列・他シート・対象セル以外は一切変更しない。
    数式・書式は openpyxl の保持機能でそのまま維持される。
    """
    wb = load_workbook(io.BytesIO(original_bytes), data_only=False, keep_vba=False)

    updated_count = 0
    skipped_count = 0
    written_log: list[dict[str, Any]] = []

    for cm in cell_maps:
        if month_from and cm.month < month_from:
            skipped_count += 1
            continue
        if month_to and cm.month > month_to:
            skipped_count += 1
            continue
        agg = monthly_summary.get(cm.month)
        if not agg:
            skipped_count += 1
            continue

        ws = wb[cm.sheet]
        value_key = "入金合計" if cm.kind == "入金" else "出金合計"
        new_value = float(agg.get(value_key, 0) or 0)
        before = ws.cell(cm.row, cm.col).value
        ws.cell(cm.row, cm.col).value = new_value
        updated_count += 1
        written_log.append(
            {
                "sheet": cm.sheet,
                "month": cm.month,
                "kind": cm.kind,
                "cell": cm.cell_address,
                "before": before,
                "after": new_value,
            }
        )

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read(), updated_count, skipped_count, written_log


def verify_overwrite(
    updated_bytes: bytes,
    cell_maps: list[BudgetCellMap],
) -> list[dict[str, Any]]:
    """更新後のファイルを再読み込みして、各セルの上書き後値を返す（比較レポート用）。"""
    wb = load_workbook(io.BytesIO(updated_bytes), data_only=False, keep_vba=False)
    out: list[dict[str, Any]] = []
    for cm in cell_maps:
        ws = wb[cm.sheet]
        after = ws.cell(cm.row, cm.col).value
        out.append(
            {
                "sheet": cm.sheet,
                "month": cm.month,
                "kind": cm.kind,
                "cell": cm.cell_address,
                "after": after,
            }
        )
    return out
